"""
Exportações:
  - gerar_pdf_memoria(): equivalente a GerarPDFMemoriaPorProjeto (VBA), um
    PDF por Projeto/SubProjeto com a memória de cálculo de diárias/passagens.
  - gerar_pdf_resumo(): PDF do resumo por Coordenadoria (aba 'Resumo').
  - gerar_excel_pca(): exporta a tabela PCA2026 para .xlsx.
  - gerar_excel_fechamento() / gerar_pdf_fechamento(): relatórios da Projeção
    de Fechamento 2026 (ver Orcamento2026.md).
"""
from pathlib import Path
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from core.config import EXPORT_DIR
from core.formatos import fmt_num_br, MESES_PT


def _styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TituloMemoria", fontSize=16, leading=20, alignment=1,
                               textColor=colors.HexColor("#002060"), spaceAfter=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Corpo", fontSize=11, leading=15))
    return styles


def gerar_pdf_memoria(indice: int, projeto: str, subprojeto: str, descricao: str, memoria: str) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORT_DIR / f"Memoria_{indice}.pdf"
    styles = _styles()
    doc = SimpleDocTemplate(str(caminho), pagesize=A4,
                             leftMargin=0.64 * cm, rightMargin=0.64 * cm,
                             topMargin=1.91 * cm, bottomMargin=1.91 * cm)
    story = [
        Paragraph("MEMÓRIA DE CÁLCULO - DIÁRIAS E PASSAGENS", styles["TituloMemoria"]),
        Spacer(1, 0.5 * cm),
        Paragraph(f"<b>Projeto:</b> {projeto}", styles["Corpo"]),
        Paragraph(f"<b>Subprojeto:</b> {subprojeto}", styles["Corpo"]),
        Spacer(1, 0.5 * cm),
        Paragraph(f"<b>Descrição:</b><br/>{descricao or '(sem descrição)'}".replace("\n", "<br/>"), styles["Corpo"]),
        Spacer(1, 0.5 * cm),
        Paragraph((memoria or "(sem memória de cálculo)").replace("\n", "<br/>"), styles["Corpo"]),
    ]
    doc.build(story)
    return caminho


def gerar_pdf_resumo(df_resumo: pd.DataFrame, coordenadoria: str, ano: int) -> Path:
    """df_resumo esperado com colunas: Projeto, Conta, Descricao, Total"""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORT_DIR / f"Resumo_{coordenadoria}_{ano}.pdf"
    styles = _styles()
    doc = SimpleDocTemplate(str(caminho), pagesize=A4)
    story = [
        Paragraph(f"Elaboração do Orçamento de {ano} - Resumo", styles["TituloMemoria"]),
        Paragraph(f"Coordenadoria: {coordenadoria}", styles["Corpo"]),
        Spacer(1, 0.5 * cm),
    ]
    data = [list(df_resumo.columns)] + df_resumo.astype(str).values.tolist()
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002060")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(tbl)
    doc.build(story)
    return caminho


def gerar_excel_pca(df_pca: pd.DataFrame) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORT_DIR / f"PCA2026_{datetime.now():%Y%m%d_%H%M}.xlsx"
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_pca.to_excel(writer, sheet_name="PCA2026", index=False)
    return caminho


# ---------------------------------------------------------------------------
# Projeção de Fechamento 2026
# ---------------------------------------------------------------------------

def _preparar_fechamento(df: pd.DataFrame):
    """Monta (resumo por natureza, detalhe por conta) a partir de
    projecao_resultado. Retorna DataFrames com valores numéricos."""
    from core.projecao_engine import NATUREZAS_ROTULOS, METODO_ROTULOS

    det = df.copy()
    det["natureza_rotulo"] = det["natureza"].map(NATUREZAS_ROTULOS).fillna(det["natureza"])
    det["metodo"] = det["metodo"].map(lambda m: METODO_ROTULOS.get(m, m))
    det["pct_exec"] = det.apply(
        lambda r: r["proj_fechamento"] / r["orcamento"] if r["orcamento"] else None, axis=1)

    ordem = [NATUREZAS_ROTULOS[n] for n in NATUREZAS_ROTULOS]
    res = (det.groupby("natureza_rotulo")
              .agg(orcamento=("orcamento", "sum"), ytd=("ytd", "sum"),
                   fechamento=("proj_fechamento", "sum"), m2=("proj_m2", "sum"),
                   n_contas=("conta", "size"))
              .reindex([n for n in ordem if n in set(det["natureza_rotulo"])])
              .reset_index())
    res["pct_exec"] = res["fechamento"] / res["orcamento"].where(res["orcamento"] != 0)

    total = pd.DataFrame([{
        "natureza_rotulo": "TOTAL", "orcamento": res["orcamento"].sum(),
        "ytd": res["ytd"].sum(), "fechamento": res["fechamento"].sum(),
        "m2": res["m2"].sum(), "n_contas": res["n_contas"].sum(),
    }])
    total["pct_exec"] = total["fechamento"] / total["orcamento"]
    res = pd.concat([res, total], ignore_index=True)

    det = det.sort_values(["natureza_rotulo", "proj_fechamento"],
                          ascending=[True, False])
    return res, det


def gerar_excel_fechamento(df: pd.DataFrame, mes_corte: int, ano: int,
                           ramo: str = "6.3",
                           df_mensal: pd.DataFrame | None = None) -> Path:
    """Excel com abas Resumo (por natureza), Detalhe (por conta, com memória de
    cálculo) e, se df_mensal for passado, Curva Mensal (conta × mês; meses
    projetados marcados com *). Valores numéricos com formato de número."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from core.projecao_engine import RAMOS_ROTULOS

    rotulo_ramo = RAMOS_ROTULOS.get(ramo, ramo)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORT_DIR / f"Fechamento{ano}_{rotulo_ramo}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    res, det = _preparar_fechamento(df)

    res_x = res.rename(columns={
        "natureza_rotulo": "Natureza", "orcamento": "Orçamento",
        "ytd": f"Realizado até {MESES_PT[mes_corte]}", "fechamento": "Fechamento Projetado (M3)",
        "pct_exec": "% Execução Projetada", "m2": "Conferência (M2)", "n_contas": "Nº Contas",
    })[["Natureza", "Nº Contas", "Orçamento", f"Realizado até {MESES_PT[mes_corte]}",
        "Fechamento Projetado (M3)", "% Execução Projetada", "Conferência (M2)"]]

    det_x = det.rename(columns={
        "conta": "Conta", "descricao": "Descrição", "natureza_rotulo": "Natureza",
        "metodo": "Método", "confianca": "Confiança", "orcamento": "Orçamento",
        "ytd": f"Realizado até {MESES_PT[mes_corte]}", "proj_restante": "Projeção Meses Restantes",
        "proj_fechamento": "Fechamento Projetado (M3)", "pct_exec": "% Execução Projetada",
        "proj_m2": "Conferência (M2)", "divergencia_pct": "Divergência M3×M2",
        "memoria": "Memória de Cálculo",
    })[["Conta", "Descrição", "Natureza", "Método", "Confiança", "Orçamento",
        f"Realizado até {MESES_PT[mes_corte]}", "Projeção Meses Restantes",
        "Fechamento Projetado (M3)", "% Execução Projetada", "Conferência (M2)",
        "Divergência M3×M2", "Memória de Cálculo"]]

    abas = [("Resumo", None), ("Detalhe", None)]
    mensal_x = None
    if df_mensal is not None and not df_mensal.empty:
        piv = df_mensal.pivot_table(index="conta", columns="mes", values="valor",
                                    aggfunc="sum", fill_value=0.0)
        piv["Total"] = piv.sum(axis=1)
        info = df_mensal.drop_duplicates("conta").set_index("conta")[["descricao", "natureza"]]
        mensal_x = info.join(piv).reset_index()
        colunas_mes = {m: MESES_PT[m] + ("*" if m > mes_corte else "") for m in range(1, 13)}
        mensal_x = mensal_x.rename(columns={"conta": "Conta", "descricao": "Descrição",
                                            "natureza": "Natureza", **colunas_mes})
        mensal_x = mensal_x.sort_values("Total", ascending=False)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        res_x.to_excel(writer, sheet_name="Resumo", index=False, startrow=2)
        det_x.to_excel(writer, sheet_name="Detalhe", index=False, startrow=2)
        if mensal_x is not None:
            mensal_x.to_excel(writer, sheet_name="Curva Mensal", index=False, startrow=2)

        titulo = (f"Projeção de Fechamento do Orçamento {ano} — CFC — {rotulo_ramo} "
                  f"(corte: {MESES_PT[mes_corte]}/{ano}, gerado em {datetime.now():%d/%m/%Y %H:%M})")
        cab_fill = PatternFill("solid", fgColor="1F4E8C")
        cab_fill_proj = PatternFill("solid", fgColor="3FA0C8")
        cab_font = Font(color="FFFFFF", bold=True)

        if mensal_x is not None:
            ws = writer.sheets["Curva Mensal"]
            ws["A1"] = titulo + " — * = mês projetado"
            ws["A1"].font = Font(bold=True, size=12, color="1F4E8C")
            for col_idx, col_nome in enumerate(mensal_x.columns, start=1):
                cel = ws.cell(row=3, column=col_idx)
                cel.fill = cab_fill_proj if str(col_nome).endswith("*") else cab_fill
                cel.font = cab_font
                letra = cel.column_letter
                if col_nome in ("Conta",):
                    ws.column_dimensions[letra].width = 19
                elif col_nome in ("Descrição",):
                    ws.column_dimensions[letra].width = 40
                elif col_nome in ("Natureza",):
                    ws.column_dimensions[letra].width = 20
                else:
                    ws.column_dimensions[letra].width = 13
                    for row in range(4, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = "#,##0.00"
            ws.freeze_panes = "C4"

        for nome, df_x in (("Resumo", res_x), ("Detalhe", det_x)):
            ws = writer.sheets[nome]
            ws["A1"] = titulo
            ws["A1"].font = Font(bold=True, size=12, color="1F4E8C")
            for cel in ws[3]:
                cel.fill = cab_fill
                cel.font = cab_font
                cel.alignment = Alignment(wrap_text=True, vertical="center")
            for col_idx, col_nome in enumerate(df_x.columns, start=1):
                letra = ws.cell(row=3, column=col_idx).column_letter
                if col_nome.startswith(("Orçamento", "Realizado", "Fechamento",
                                        "Conferência", "Projeção")):
                    largura = 18
                    fmt = "#,##0.00"
                elif col_nome.startswith(("% ", "Divergência")):
                    largura = 12
                    fmt = "0.0%"
                elif col_nome in ("Descrição", "Memória de Cálculo"):
                    largura = 60 if col_nome == "Memória de Cálculo" else 42
                    fmt = None
                elif col_nome == "Conta":
                    largura = 19
                    fmt = None
                else:
                    largura = 14
                    fmt = None
                ws.column_dimensions[letra].width = largura
                if fmt:
                    for row in range(4, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = fmt
            # negrito na linha TOTAL do Resumo
            if nome == "Resumo":
                for cel in ws[ws.max_row]:
                    cel.font = Font(bold=True)
            ws.freeze_panes = "A4"
    return caminho


def gerar_pdf_fechamento(df: pd.DataFrame, mes_corte: int, ano: int,
                         ramo: str = "6.3",
                         df_mensal: pd.DataFrame | None = None) -> Path:
    """PDF (A4 paisagem): resumo por natureza, curva mensal (se df_mensal),
    detalhe por conta e alertas."""
    from core.projecao_engine import LIMIAR_DIVERGENCIA, RAMOS_ROTULOS, NATUREZAS_ROTULOS

    rotulo_ramo = RAMOS_ROTULOS.get(ramo, ramo)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORT_DIR / f"Fechamento{ano}_{rotulo_ramo}_{datetime.now():%Y%m%d_%H%M}.pdf"
    res, det = _preparar_fechamento(df)
    styles = _styles()

    def num(v):
        return fmt_num_br(v)

    def pct(v):
        return f"{v * 100:,.1f}%".replace(",", "_").replace(".", ",").replace("_", ".") if v == v and v is not None else "—"

    doc = SimpleDocTemplate(str(caminho), pagesize=landscape(A4),
                            leftMargin=1 * cm, rightMargin=1 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    story = [
        Paragraph(f"Projeção de Fechamento do Orçamento {ano} — CFC — {rotulo_ramo}",
                  styles["TituloMemoria"]),
        Paragraph(f"Corte: {MESES_PT[mes_corte]}/{ano} · Gerado em {datetime.now():%d/%m/%Y %H:%M} · "
                  "Modelo M3 (híbrido por natureza) com conferência M2 (sazonal)", styles["Corpo"]),
        Spacer(1, 0.4 * cm),
    ]

    estilo_tab = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e8c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dfe6ee")),
    ])

    # --- Resumo por natureza ---
    cab = ["Natureza", "Contas", "Orçamento", f"Realizado até {MESES_PT[mes_corte]}",
           "Fechamento (M3)", "% Exec", "Conferência (M2)"]
    linhas = [cab] + [[r["natureza_rotulo"], int(r["n_contas"]), num(r["orcamento"]),
                       num(r["ytd"]), num(r["fechamento"]), pct(r["pct_exec"]),
                       num(r["m2"])] for _, r in res.iterrows()]
    tbl = Table(linhas, repeatRows=1)
    tbl.setStyle(estilo_tab)
    story += [Paragraph(f"Resumo por natureza de {rotulo_ramo.lower()}", styles["Heading3"]), tbl,
              Spacer(1, 0.5 * cm)]

    # --- Curva mensal por natureza (R$ mil) ---
    if df_mensal is not None and not df_mensal.empty:
        def mil(v):
            return f"{v / 1000:,.0f}".replace(",", ".")

        cab_m = ["Natureza"] + [MESES_PT[m] + ("*" if m > mes_corte else "") for m in range(1, 13)] + ["Total"]
        linhas_m = [cab_m]
        ordem_nats = [n for n in NATUREZAS_ROTULOS if n in set(df_mensal["natureza"])]
        for nat in ordem_nats:
            sub_m = df_mensal[df_mensal["natureza"] == nat]
            por_mes = sub_m.groupby("mes")["valor"].sum()
            linhas_m.append([NATUREZAS_ROTULOS[nat][:28]]
                            + [mil(por_mes.get(m, 0.0)) for m in range(1, 13)]
                            + [mil(sub_m["valor"].sum())])
        por_mes_tot = df_mensal.groupby("mes")["valor"].sum()
        linhas_m.append(["TOTAL"] + [mil(por_mes_tot.get(m, 0.0)) for m in range(1, 13)]
                        + [mil(df_mensal["valor"].sum())])
        tbl_m = Table(linhas_m, repeatRows=1)
        tbl_m.setStyle(estilo_tab)
        story += [Paragraph("Curva mensal por natureza — valores em R$ mil "
                            "(* = mês projetado)", styles["Heading3"]), tbl_m,
                  Spacer(1, 0.5 * cm)]

    # --- Alertas ---
    det["material"] = det[["proj_fechamento", "proj_m2"]].abs().max(axis=1) > 50_000
    divergentes = det[(det["divergencia_pct"] > LIMIAR_DIVERGENCIA) & det["material"]]
    baixa = det[(det["confianca"] == "baixa") & (det["orcamento"] > 0)]
    if not divergentes.empty or not baixa.empty:
        alertas = []
        if not divergentes.empty:
            alertas.append(f"{len(divergentes)} conta(s) com divergência M3×M2 acima de "
                           f"{LIMIAR_DIVERGENCIA:.0%} (destacadas no detalhe com *).")
        if not baixa.empty:
            alertas.append(f"{len(baixa)} conta(s) com confiança BAIXA — revisar ou "
                           "definir Fechamento Manual com cronograma (destacadas com ‡).")
        story += [Paragraph("Pontos de atenção: " + " ".join(alertas), styles["Corpo"]),
                  Spacer(1, 0.4 * cm)]

    # --- Detalhe por conta ---
    cab_det = ["Conta", "Descrição", "Método", "Orçamento", "Realizado",
               "Fechamento (M3)", "% Exec", "M2", "Diverg."]
    linhas_det = [cab_det]
    marca_div = set(divergentes["conta"])
    marca_baixa = set(baixa["conta"])
    for _, r in det.iterrows():
        marcas = ("*" if r["conta"] in marca_div else "") + ("‡" if r["conta"] in marca_baixa else "")
        linhas_det.append([
            r["conta"] + marcas, str(r["descricao"])[:44], r["metodo"],
            num(r["orcamento"]), num(r["ytd"]), num(r["proj_fechamento"]),
            pct(r["pct_exec"]), num(r["proj_m2"]), pct(r["divergencia_pct"]),
        ])
    linhas_det.append(["TOTAL", "", "", num(det["orcamento"].sum()), num(det["ytd"].sum()),
                       num(det["proj_fechamento"].sum()), pct(det["proj_fechamento"].sum() / det["orcamento"].sum()),
                       num(det["proj_m2"].sum()), ""])
    tbl_det = Table(linhas_det, repeatRows=1)
    tbl_det.setStyle(estilo_tab)
    story += [Paragraph(f"Detalhe por conta (contas folha {ramo})", styles["Heading3"]), tbl_det,
              Spacer(1, 0.3 * cm),
              Paragraph("* divergência material M3×M2 acima do limiar · ‡ confiança baixa. "
                        "A memória de cálculo de cada conta está disponível no Excel e na tela "
                        "Fechamento 2026 do sistema.", styles["Corpo"])]

    doc.build(story)
    return caminho
